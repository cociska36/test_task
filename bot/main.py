import logging
import time
from aiogram import Bot, Dispatcher, types, Router
from aiogram.filters import CommandStart
import asyncio
from dotenv import load_dotenv
from database import Database
from config import settings
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery, InlineQuery, InlineQueryResultArticle, InputTextMessageContent
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from aiogram import types
from aiogram.types import InlineQueryResultArticle, InputTextMessageContent
from yookassa import Configuration, Payment
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup


router = Router()
load_dotenv()

TELEGRAM_API_TOKEN = settings.bot_token 
DB_HOST = settings.postgres_host 
DB_PORT = settings.postgres_port  
DB_NAME = settings.postgres_db 
DB_USER = settings.postgres_user
DB_PASSWORD = settings.postgres_password 
API_TOKEN = settings.bot_token

CHANNEL_ID = settings.channel_id
GROUP_ID = settings.group_id 

# Настраиваем идентификатор магазина и секретный ключ
Configuration.account_id = 'your_shop_id'  # Ваш Shop ID
Configuration.secret_key = 'your_secret_key'

# Логирование
logging.basicConfig(level=logging.INFO)

bot = Bot(token=TELEGRAM_API_TOKEN)
dispatcher = Dispatcher()
logging.info(f"Connecting to database with settings: "
             f"host={DB_HOST}, port={DB_PORT}, dbname={DB_NAME}, user={DB_USER}")

# Инициализация базы данных
db = Database(DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD, bot)

class OrderStates(StatesGroup):
    waiting_for_delivery_data = State()  # Ожидание данных для доставки


# Пример данных о категориях, подкатегориях и товарах
CATEGORY_PAGE_SIZE = 3  # Количество категорий на странице
categories = [
    {'id': 1, 'name': 'Электроника'},
    {'id': 2, 'name': 'Одежда'},
    {'id': 3, 'name': 'Книги'},
    {'id': 4, 'name': 'Косметика'},
    {'id': 5, 'name': 'Игры'},
    {'id': 6, 'name': 'Мебель'},
]
subcategories = {
    1: [{'id': 1, 'name': 'Телевизоры'}, {'id': 2, 'name': 'Смартфоны'}, {'id': 3, 'name': 'Ноутбуки'}],  # Электроника
    2: [{'id': 4, 'name': 'Мужская одежда'}, {'id': 5, 'name': 'Женская одежда'}],  # Одежда
}
products = {
    1: [
        {
            'id': 1,
            'name': 'Товар 1',
            'description': 'Описание товара 1',
            'image_url': 'https://example.com/images/product1.jpg'  # URL изображения (необязательно)
        },
        {
            'id': 2,
            'name': 'Товар 2',
            'description': 'Описание товара 2',
            'image_url': 'https://example.com/images/product2.jpg'  # URL изображения (необязательно)
        }
    ]
}

# Данные для FAQ
faq_data = [
    {"question": "Как сделать заказ?", "answer": "Для оформления заказа выберите товар и следуйте инструкциям."},
    {"question": "Какие способы оплаты доступны?", "answer": "Мы принимаем оплату картой и через электронные кошельки."},
    {"question": "Как изменить адрес доставки?", "answer": "Свяжитесь с поддержкой для изменения адреса."},
    {"question": "Как отслеживать статус заказа?", "answer": "Вы можете отслеживать статус через нашу систему трекинга."},
]


# Хранение состояния корзины и текущих страниц
category_pages = {}
cart = {}

# Функция для проверки подписки на канал
async def check_subscription(user_id: int, chat_id: str) -> bool:
    try:
        chat_member = await bot.get_chat_member(chat_id, user_id)
        if chat_member.status in ['member', 'administrator', 'creator']:
            return True
        else:
            return False
    except Exception as e:
        logging.error(f"Ошибка при проверке подписки пользователя {user_id} на {chat_id}: {e}")
        return False

# Обработчик команды /start
@dispatcher.message(CommandStart())
async def cmd_start(message: types.Message):
    user_id = message.from_user.id

    # Проверяем подписку на канал
    is_subscribed_to_channel = await check_subscription(user_id, CHANNEL_ID)
    # Проверяем участие в группе
    is_member_of_group = await check_subscription(user_id, GROUP_ID)
    
    if is_subscribed_to_channel and is_member_of_group:
        
        # Создаем инлайн-клавиатуру с кнопками
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Каталог", callback_data="catalog"),
             InlineKeyboardButton(text="Корзина", callback_data="cart"),
             InlineKeyboardButton(text="FAQ", callback_data="faq")]
        ])  
        await message.answer("Привет, ты подписан на наш канал и состоишь в группе! Добро пожаловать! "
                             "Выбери одну из опций:", reply_markup=keyboard)
        
        # Добавляем пользователя в базу данных
        await db.add_user(user_id)
    else:
        logging.warning(f"Пользователь {user_id} не прошел проверку подписки и участия.")
        await message.answer(
            "Для использования бота необходимо подписаться на канал и стать участником группы."
            "\nПожалуйста, подпишитесь на канал и вступите в группу."
        )

# Функция для отображения страницы категорий
async def send_category_page(message, page=1):
    
    # Вычисляем индекс для выбора категорий на текущей странице
    start_index = (page - 1) * CATEGORY_PAGE_SIZE
    end_index = start_index + CATEGORY_PAGE_SIZE
    page_categories = categories[start_index:end_index]
    
    try:
        # Создаем кнопки для каждой категории на текущей странице
        buttons = [
            InlineKeyboardButton(text=category['name'], callback_data=f"category_{category['id']}")
            for category in page_categories
        ]
    except Exception as e:
        logging.error(f"Ошибка при создании кнопок: {e}")
        raise
    
    # Создаем кнопки для навигации по страницам
    pagination_buttons = []
    if page > 1:
        pagination_buttons.append(InlineKeyboardButton(text="« Предыдущая", callback_data=f"category_page_{page - 1}"))
    if end_index < len(categories):
        pagination_buttons.append(InlineKeyboardButton(text="Следующая »", callback_data=f"category_page_{page + 1}"))

    # Формируем клавиатуру, добавляя кнопки категории и пагинации (если есть)
    keyboard = InlineKeyboardMarkup(inline_keyboard=[buttons] + [pagination_buttons] if pagination_buttons else [])
    
    # Отправляем сообщение с клавиатурой
    await message.answer("Выберите категорию:", reply_markup=keyboard)


@dispatcher.callback_query(lambda c: c.data == 'catalog')
async def handle_catalog(call: CallbackQuery):
    user_id = call.from_user.id
    page = category_pages.get(user_id, 1)
    await send_category_page(call.message, page)

@dispatcher.callback_query(lambda c: c.data.startswith('category_page_'))
async def handle_category_page(call: CallbackQuery):
    user_id = call.from_user.id
    page = int(call.data.split('_')[-1])
    category_pages[user_id] = page
    await send_category_page(call.message, page)

@dispatcher.callback_query(lambda c: c.data.startswith('category_'))
@dispatcher.callback_query(lambda c: c.data.startswith('category_'))
async def handle_category_selection(call: CallbackQuery):
    # Извлекаем ID категории из данных callback
    category_id = int(call.data.split('_')[1])
    
    # Получаем подкатегории для выбранной категории из словаря subcategories
    category_subcategories = subcategories.get(category_id, [])
    
    buttons = []
    
    # Создание кнопок для подкатегорий
    for subcategory in category_subcategories:
        button = InlineKeyboardButton(
            text=subcategory['name'],  # Убедитесь, что 'name' является строкой
            callback_data=f"subcategory_{subcategory['id']}"  # Убедитесь, что 'id' является корректным значением
        )
        buttons.append(button)
    
    # Создаем клавиатуру с кнопками подкатегорий
    keyboard = InlineKeyboardMarkup(inline_keyboard=[buttons])
    
    # Отправляем сообщение с клавиатурой
    await call.message.answer("Выберите подкатегорию:", reply_markup=keyboard)


@dispatcher.callback_query(lambda c: c.data.startswith('subcategory_'))
async def handle_subcategory_selection(call: CallbackQuery):
    subcategory_id = int(call.data.split('_')[1])
    
    product_list = products.get(subcategory_id, [])
    
    buttons = []
    for product in product_list:
        button = InlineKeyboardButton(text=product['name'], callback_data=f"product_{product['id']}")
        buttons.append(button)

    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[buttons])
    await call.message.answer("Выберите товар:", reply_markup=keyboard)

@dispatcher.callback_query(lambda c: c.data.startswith('product_'))
async def handle_product_selection(call: CallbackQuery):
    try:
        # Разделяем callback_data для получения product_id
        parts = call.data.split('_')
        if len(parts) < 2:
            raise ValueError(f"Unexpected callback data format: {call.data}")
        
        product_id = int(parts[1])
        
        product = next((p for p in products.get(1, []) if p['id'] == product_id), None)
        
        if product is None:
            raise ValueError(f"Product with id {product_id} not found.")
        
        user_id = call.from_user.id
        cart[user_id] = cart.get(user_id, [])
        cart[user_id].append({'product_id': product_id, 'quantity': 1})

        # Формируем сообщение с данными о товаре (описание и изображение)
        product_name = product['name']
        product_description = product['description']
        product_image_url = product.get('image_url', None)

        message_text = f"Выбран товар: {product_name}\nОписание: {product_description}\nУкажите количество:"
        
        logging.info(product_image_url)
        if product_image_url:
            await call.message.answer(message_text, 
                                      photo=product_image_url)
        else:
            await call.message.answer(message_text)

    except ValueError as e:
        logging.error(f"Ошибка в формате callback_data: {e}")
        await call.message.answer("Ошибка при обработке выбора товара. Попробуйте снова.")



# Обработчик для сообщений, где текст является числом (например, количество товара)
@dispatcher.message(lambda message: message.text.isdigit())
async def handle_quantity(message: types.Message):
    user_id = message.from_user.id
    quantity = int(message.text)

    if cart.get(user_id):
        cart[user_id][-1]['quantity'] = quantity
        confirm_button = InlineKeyboardButton(text="Подтвердить", callback_data="confirm_order")
        
        # Создаем клавиатуру с кнопкой подтверждения
        keyboard = InlineKeyboardMarkup(inline_keyboard=[[confirm_button]])
        
        await message.answer("Товар добавлен в корзину! Подтвердите покупку.", reply_markup=keyboard)
    else:
        logging.error(f"Ошибка: пользователь {user_id} попытался подтвердить покупку, но корзина пуста.")

# Обработчики для кнопок "Корзина" и "FAQ"
@dispatcher.callback_query(lambda c: c.data == 'cart')
async def handle_cart(call: CallbackQuery):
    user_id = call.from_user.id
    
    # Проверяем, есть ли товары в корзине
    if cart.get(user_id):
        basket_info = ""
        
        # Заменяем ID на имя товара
        for item in cart[user_id]:
            product_id = item['product_id']
            # Получаем продукт по ID
            product = next((p for p in products.get(1, []) if p['id'] == product_id), None)
            if product:
                basket_info += f"{product['name']} x{item['quantity']}\n"  # Используем name вместо id
        
        # Создаем кнопки для удаления товаров
        remove_buttons = [
            InlineKeyboardButton(text=f"Удалить {item['product_id']}", callback_data=f"remove_{item['product_id']}")
            for item in cart[user_id]
        ]
        
        # Добавляем кнопки удаления и кнопку перехода к следующему шагу
        remove_buttons.append(InlineKeyboardButton(text="Перейти к оформлению", callback_data="checkout"))
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=[remove_buttons])
        await call.message.answer(f"Ваша корзина:\n{basket_info}", reply_markup=keyboard)
    else:
        await call.answer("Ваша корзина пуста!")



@dispatcher.callback_query(lambda c: c.data == 'confirm_order')
async def handle_confirm_order(call: CallbackQuery):
    user_id = call.from_user.id
    
    # Перенаправляем пользователя в корзину
    await handle_cart(call)

@dispatcher.callback_query(lambda c: c.data.startswith('remove_'))
async def handle_remove_item(call: CallbackQuery):
    user_id = call.from_user.id
    product_id = int(call.data.split('_')[1])
    
    # Удаляем товар из корзины
    cart[user_id] = [item for item in cart[user_id] if item['product_id'] != product_id]
    
    await call.answer(f"Товар {product_id} удален из корзины!")
    await handle_cart(call)

@dispatcher.callback_query(lambda c: c.data == 'checkout')
async def handle_checkout(call: CallbackQuery, state: FSMContext):
    user_id = call.from_user.id
    
    # Запрашиваем данные для доставки
    await call.message.answer("Пожалуйста, укажите ваши данные для доставки (имя, адрес, телефон):")
    
    # Сохраняем заказ в базе данных
    await db.save_order(user_id, cart[user_id])  # Сохранение заказа
    
    # Переводим пользователя в состояние ожидания данных для доставки
    await state.set_state(OrderStates.waiting_for_delivery_data)


# Обработчик для получения данных о доставке
@dispatcher.message(lambda message: message.text and not message.text.lower().startswith('@'))
async def handle_delivery_data(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    
    # Проверяем, находится ли пользователь в нужном состоянии
    current_state = await state.get_state()
    if current_state != OrderStates.waiting_for_delivery_data.state:
        return  # Игнорируем сообщение, если пользователь не в нужном состоянии

    # Проверяем, есть ли у пользователя корзина
    if user_id not in cart or not cart[user_id]:
        await message.answer("Корзина пуста. Пожалуйста, добавьте товары в корзину.")
        return

    delivery_data = message.text

    # Сохраняем данные для доставки в базу данных
    await db.save_delivery_data(user_id, delivery_data)

    # Сохраняем заказ в БД
    await db.save_order(user_id, cart.get(user_id, []))

    # После сохранения заказа, создаем кнопку и отправляем сообщение
    remove_buttons = [
        InlineKeyboardButton(text="Оплатить", callback_data="pay_")
    ]
    keyboard = InlineKeyboardMarkup(inline_keyboard=[remove_buttons])
    await message.answer("Ваш заказ успешно создан! Нажмите на кнопку ниже для оплаты.", reply_markup=keyboard)
    
    # Заканчиваем процесс оформления заказа (можно сбросить состояние, если нужно)
    await state.clear()  # Очищаем состояние

    
# Функция для получения последнего order_id
def get_last_order_id():
    try:
        # Открытие или создание файла Excel
        try:
            wb = openpyxl.load_workbook('orders.xlsx')  # Попытка загрузить существующий файл
        except FileNotFoundError:
            wb = Workbook()  # Если файл не существует, создаем новый
            sheet = wb.active
            sheet.title = "Orders"
            sheet.append(['Order ID', 'User ID', 'Product ID', 'Quantity', 'Amount', 'Order Date'])
            wb.save('orders.xlsx')
            return 0  # Если нет данных, начинаем с 0

        # Получаем активный лист
        sheet = wb.active

        # Получаем последний номер order_id
        last_row = sheet.max_row
        if last_row == 1:  # если только заголовки в таблице
            return 0
        last_order_id = sheet.cell(row=last_row, column=1).value
        return last_order_id
    except Exception as e:
        logging.error(f"Ошибка при получении последнего order_id: {e}")
        return 0

# Функция для сохранения данных заказа в Excel
async def save_order_to_excel(user_id, cart, amount):
    try:
        # Получаем последний order_id и увеличиваем его на 1
        last_order_id = get_last_order_id()
        order_id = last_order_id + 1

        # Открытие или создание файла Excel
        try:
            wb = openpyxl.load_workbook('orders.xlsx')  # Попытка загрузить существующий файл
        except FileNotFoundError:
            wb = Workbook()  # Создаем новый файл, если его нет
            sheet = wb.active
            sheet.title = "Orders"
            sheet.append(['Order ID', 'User ID', 'Product ID', 'Quantity', 'Amount', 'Order Date'])

        # Получаем активный лист
        sheet = wb.active

        # Добавляем данные заказа в Excel
        for item in cart:
            product_id = item['product_id']
            quantity = item['quantity']
            sheet.append([order_id, user_id, product_id, quantity, amount, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])

        # Сохраняем файл Excel
        wb.save('orders.xlsx')
        logging.info(f"Заказ {order_id} добавлен в Excel.")
        return order_id
    except Exception as e:
        logging.error(f"Ошибка при сохранении заказа в Excel: {e}")
        return None

# Обработчик успешного платежа
@dispatcher.callback_query(lambda c: c.data.startswith('pay_'))
async def handle_payment(callback_query: types.CallbackQuery):
    user_id = callback_query.from_user.id

    # Тестовые данные для извлечения order_id
    data = ['pay', '100']  # Заданные данные для тестирования
    logging.info(data)
    if len(data) < 2 or not data[1].isdigit():
        await callback_query.message.answer("Некорректные данные для оплаты.")
        return

    try:
        amount = float(data[1]) 

        # Создаем платеж через Юкассу
        payment = Payment.create({
            "amount": {
                "value": str(amount),
                "currency": "RUB"
            },
            "confirmation": {
                "type": "redirect",
                "return_url": "https://example.com/success"
            },
            "capture": True,
            "description": f"Оплата заказа пользователем {user_id}"
        })

        # Получаем ссылку на оплату
        payment_url = payment.confirmation.confirmation_url

        # Отправляем ссылку пользователю
        await callback_query.message.answer(
            f"Для оплаты вашего заказа перейдите по ссылке: {payment_url}\nСумма: {amount} ₽"
        )
        time.sleep(30)
        payment_status = Payment.find_one(payment.id).status
        if payment_status == 'succeeded':
            cart_items = cart.get(user_id, [])
            order_amount = 100  # Примерная сумма заказа, нужно извлечь из контекста платежа
            order_id = await save_order_to_excel(user_id, cart_items, order_amount)
            if order_id:
                await callback_query.message.answer(f"Оплата успешно завершена! Ваш заказ ID {order_id} сохранен в Excel.")
            else:
                await callback_query.message.answer("Произошла ошибка при сохранении заказа в Excel.")
        else:
            await callback_query.message.answer("Платеж не был завершен. Попробуйте еще раз.")
    except Exception as e:
        # Обрабатываем возможные ошибки
        await callback_query.message.answer(
            f"Произошла ошибка при создании платежа: {e}"
        )

# Функция для сохранения заказа
async def save_order(self, user_id, cart):
    try:
        # Предполагается, что у вас есть таблица 'orders' для хранения заказов
        order_date = datetime.now()
         
        async with self.pool.acquire() as connection:
            # Вставляем заказ в таблицу заказов
            order_id = await connection.fetchval("""
                INSERT INTO orders (user_id, order_date, status)
                VALUES ($1, $2, $3) RETURNING order_id;
            """, user_id, order_date, 'new')
             
            # Вставляем каждый товар из корзины в таблицу order_items
            for item in cart:
                await connection.execute("""
                    INSERT INTO order_items (order_id, product_id, quantity)
                    VALUES ($1, $2, $3);
                """, order_id, item['product_id'], item['quantity'])
            
            logging.info(f"Заказ сохранён для пользователя {user_id} с order_id {order_id}")
            
            # Отправляем пользователю кнопку "Оплатить"
            markup = InlineKeyboardMarkup().add(
                InlineKeyboardButton("Оплатить", callback_data=f'pay_{order_id}')
            )
            # Отправляем сообщение пользователю
            await bot.send_message(user_id, "Ваш заказ успешно создан! Нажмите на кнопку ниже для оплаты.", reply_markup=markup)
            
    except Exception as e:
        logging.error(f"Ошибка при сохранении заказа для пользователя {user_id}: {e}")


@dispatcher.callback_query(lambda c: c.data == 'checkout')
async def handle_checkout(call: CallbackQuery):
    user_id = call.from_user.id
    
    # Запрашиваем данные для доставки
    await call.message.answer("Пожалуйста, укажите ваши данные для доставки (имя, адрес, телефон):")
    
    # Сохраняем заказ и после этого отправляем кнопку оплаты
    await db.save_order(user_id, cart[user_id])  # Сохранение заказа
    
    # После сохранения заказа, создаем кнопку и отправляем сообщение
    await call.message.answer("Ваш заказ успешно создан! Нажмите на кнопку ниже для оплаты.")


@dispatcher.callback_query(lambda c: c.data == 'faq')
async def handle_faq_button(call: types.CallbackQuery):
    """Обработка нажатия кнопки FAQ."""
    await call.message.answer(
        "Введите ключевые слова вашего вопроса в строку поиска в формате inline (например, @bot_username ваш вопрос)."
    )

# Обработчик inline-запросов для FAQ
@dispatcher.inline_query()
async def handle_inline_faq_query(inline_query: types.InlineQuery):
    query = inline_query.query.lower()  # Получаем запрос пользователя
    results = []

    # Если запрос пустой, ничего не ищем
    if not query:
        await inline_query.answer(results)
        return

    # Ищем совпадения по ключевым словам в вопросах FAQ
    for index, faq in enumerate(faq_data):
        if query in faq["question"].lower():
            results.append(
                InlineQueryResultArticle(
                    id=str(index),
                    title=faq["question"],
                    input_message_content=InputTextMessageContent(
                        message_text=f"**{faq['question']}**\n\n{faq['answer']}",  # Используем именованные параметры
                        parse_mode="Markdown"  # Указываем, что ответ должен быть в Markdown
                    ),
                    description=faq["answer"][:50],  # Описание, сокращенное до первых 50 символов
                )
            )

    # Если результатов нет, выводим сообщение "Ничего не найдено"
    if not results:
        results.append(
            InlineQueryResultArticle(
                id="no_results",
                title="Ничего не найдено",
                input_message_content=InputTextMessageContent(
                    message_text="К сожалению, по вашему запросу ничего не найдено. Попробуйте изменить запрос."
                ),
                description="Попробуйте изменить запрос.",
            )
        )

    # Отправляем результаты пользователю
    await inline_query.answer(results, cache_time=1, is_personal=True)

async def on_start():
    try:
        logging.info("Bot started!")
        await db.init()
    except Exception as e:
        logging.error(f"Ошибка инициализации базы данных: {e}")
        raise e 

if __name__ == "__main__":
    try:
        loop = asyncio.get_event_loop()
        loop.run_until_complete(on_start())  # Инициализация базы данных
        loop.run_until_complete(dispatcher.start_polling(bot))  # Запуск бота
    except Exception as e:
        logging.error(f"Ошибка при запуске бота: {e}")

